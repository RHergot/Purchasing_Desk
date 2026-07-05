"""CSV and Excel export utilities for Purchasing Desk KPIs.

Provides reusable export functions that any controller can call
to dump table data to CSV or Excel files.
"""

import csv
import logging
import os
from datetime import datetime
from decimal import Decimal
from typing import List, Optional

log = logging.getLogger(__name__)


def _model_to_rows(model) -> List[List[str]]:
    """Extract all data from a QStandardItemModel as a list of rows.

    Row 0 = header labels; rows 1..N = data.
    """
    if model is None:
        return []

    headers = []
    for col in range(model.columnCount()):
        item = model.horizontalHeaderItem(col)
        headers.append(item.text() if item else f"Col{col}")

    rows = [headers]
    for r in range(model.rowCount()):
        row = []
        for c in range(model.columnCount()):
            item = model.item(r, c)
            if item is None:
                row.append("")
            elif hasattr(item, "text"):
                row.append(item.text())
            else:
                row.append(str(item))
        rows.append(row)

    return rows


def export_csv(model, file_path: str, encoding: str = "utf-8-sig") -> bool:
    """Export a QStandardItemModel to a CSV file.

    Args:
        model: QStandardItemModel with headers and data.
        file_path: Destination file path (should end in .csv).
        encoding: File encoding (default utf-8-sig for Excel compatibility).

    Returns:
        True on success, False on failure.
    """
    try:
        rows = _model_to_rows(model)
        if not rows:
            log.warning("export_csv: model has no data")
            return False

        with open(file_path, "w", newline="", encoding=encoding) as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        log.info(f"CSV exported to {file_path} ({len(rows) - 1} rows)")
        return True
    except Exception:
        log.exception("export_csv failed:")
        return False


def export_excel(models: List, file_path: str) -> bool:
    """Export multiple QStandardItemModels to an Excel workbook.

    Each model becomes a separate sheet named by its table's header label
    or a fallback name.

    Args:
        models: List of (model, sheet_name) tuples.
        file_path: Destination file path (should end in .xlsx).

    Returns:
        True on success, False on failure.
    """
    try:
        import openpyxl
    except ImportError:
        log.error("export_excel: openpyxl is not installed. Install with: pip install openpyxl")
        return False

    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for idx, (model, sheet_name) in enumerate(models):
            rows = _model_to_rows(model)
            if not rows:
                continue

            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit: 31 chars

            for row_data in rows:
                ws.append(row_data)

            # Auto-fit column widths (approximate)
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        if not wb.sheetnames:
            log.warning("export_excel: no sheets to write")
            return False

        wb.save(file_path)
        log.info(f"Excel exported to {file_path} ({len(wb.sheetnames)} sheets)")
        return True
    except Exception:
        log.exception("export_excel failed:")
        return False


def generate_export_filename(prefix: str, ext: str) -> str:
    """Generate a timestamped export filename.

    Example: generate_export_filename("orders_kpi", ".csv")
    → "orders_kpi_20260705_120000.csv"
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}{ext}"


def export_orders_kpi(view) -> Optional[str]:
    """Export all Orders KPI data to a single CSV.

    Returns the file path on success, None on failure.
    """
    file_path = os.path.join(
        os.path.expanduser("~"), "Documents",
        generate_export_filename("orders_kpi", ".csv")
    )
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Collect models from the view
    models_exported = 0
    for attr_name in ["top_suppliers_table", "top_pieces_table"]:
        table = getattr(view, attr_name, None)
        if table and table.model():
            if export_csv(table.model(), file_path):
                models_exported += 1

    return file_path if models_exported > 0 else None


def export_financial_kpi(view) -> Optional[str]:
    """Export all Financial KPI data to a single CSV.

    Returns the file path on success, None on failure.
    """
    file_path = os.path.join(
        os.path.expanduser("~"), "Documents",
        generate_export_filename("financial_kpi", ".csv")
    )
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    models_exported = 0
    for attr_name in ["savings_table", "price_trend_table"]:
        table = getattr(view, attr_name, None)
        if table and table.model():
            if export_csv(table.model(), file_path):
                models_exported += 1

    return file_path if models_exported > 0 else None


def export_negotiations(view) -> Optional[str]:
    """Export the Negotiations KPI table to a CSV.

    Returns the file path on success, None on failure.
    """
    file_path = os.path.join(
        os.path.expanduser("~"), "Documents",
        generate_export_filename("negotiations_kpi", ".csv")
    )
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    table = getattr(view, "table_view", None)
    if table and table.model():
        if export_csv(table.model(), file_path):
            return file_path

    return None
